from __future__ import annotations

import argparse
import json
import math
import shutil
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATES = PROJECT_ROOT / "problem-statements" / "attachments"
OUTPUT = PROJECT_ROOT / "results" / "working"
DRONE_INITIALS = {
    "FY1": (17800.0, 0.0, 1800.0),
    "FY2": (12000.0, 1400.0, 1400.0),
    "FY3": (6000.0, -3000.0, 700.0),
    "FY4": (11000.0, 2000.0, 1800.0),
    "FY5": (13000.0, -2000.0, 1300.0),
}


def duration(intervals: list[list[float]]) -> float:
    return sum(right - left for left, right in intervals)


def assert_row(actual: list, expected: list) -> None:
    for actual_value, expected_value in zip(actual, expected, strict=True):
        if isinstance(expected_value, (int, float)) and not isinstance(expected_value, bool):
            if not math.isclose(actual_value, expected_value, rel_tol=1e-12, abs_tol=1e-10):
                raise AssertionError(f"numeric XLSX mismatch: {actual_value} != {expected_value}")
        elif actual_value != expected_value:
            raise AssertionError(f"XLSX mismatch: {actual_value!r} != {expected_value!r}")


def validate_q5_constraints(records: list[dict]) -> None:
    by_drone: dict[str, list[dict]] = {}
    for record in records:
        drone = record["drone"]
        by_drone.setdefault(drone, []).append(record)
        if drone not in DRONE_INITIALS or record["assigned_missile"] not in {"M1", "M2", "M3"}:
            raise AssertionError("invalid drone or missile assignment")
        if not 70.0 <= record["speed_m_s"] <= 140.0:
            raise AssertionError(f"{drone}: speed outside [70, 140] m/s")
        if record["release_time_s"] < 0.0 or record["fuse_delay_s"] < 0.0:
            raise AssertionError(f"{drone}: negative release time or fuse delay")
        if not math.isclose(
            record["release_time_s"] + record["fuse_delay_s"],
            record["explosion_time_s"],
            abs_tol=1e-10,
        ):
            raise AssertionError(f"{drone}: inconsistent explosion time")

        heading = math.radians(record["heading_deg"])
        velocity = (
            record["speed_m_s"] * math.cos(heading),
            record["speed_m_s"] * math.sin(heading),
        )
        initial = DRONE_INITIALS[drone]
        expected_release = (
            initial[0] + velocity[0] * record["release_time_s"],
            initial[1] + velocity[1] * record["release_time_s"],
            initial[2],
        )
        expected_explosion = (
            record["release_point_m"][0] + velocity[0] * record["fuse_delay_s"],
            record["release_point_m"][1] + velocity[1] * record["fuse_delay_s"],
            record["release_point_m"][2] - 4.9 * record["fuse_delay_s"] ** 2,
        )
        for actual, expected in zip(record["release_point_m"], expected_release, strict=True):
            if not math.isclose(actual, expected, rel_tol=1e-12, abs_tol=1e-8):
                raise AssertionError(f"{drone}: release point violates shared trajectory")
        for actual, expected in zip(record["explosion_point_m"], expected_explosion, strict=True):
            if not math.isclose(actual, expected, rel_tol=1e-12, abs_tol=1e-8):
                raise AssertionError(f"{drone}: explosion point violates projectile motion")
        if record["explosion_point_m"][2] < 0.0:
            raise AssertionError(f"{drone}: explosion below ground")

    for drone, drone_records in by_drone.items():
        if len(drone_records) > 3 or len({record["bomb"] for record in drone_records}) != len(drone_records):
            raise AssertionError(f"{drone}: invalid bomb count or duplicate bomb number")
        headings = {record["heading_deg"] for record in drone_records}
        speeds = {record["speed_m_s"] for record in drone_records}
        if len(headings) != 1 or len(speeds) != 1:
            raise AssertionError(f"{drone}: bombs do not share one trajectory")
        releases = sorted(record["release_time_s"] for record in drone_records)
        if any(right - left < 1.0 - 1e-10 for left, right in zip(releases, releases[1:])):
            raise AssertionError(f"{drone}: adjacent releases are less than 1 s apart")


def copy_template(name: str):
    target = OUTPUT / name
    shutil.copyfile(TEMPLATES / name, target)
    workbook = openpyxl.load_workbook(target)
    return target, workbook, workbook.active


def write_q3() -> None:
    data = json.loads((OUTPUT / "q3.json").read_text(encoding="utf-8"))
    target, workbook, sheet = copy_template("result1.xlsx")
    for row, bomb in enumerate(data["bombs"], 2):
        values = [
            data["heading_deg"], data["speed_m_s"], bomb["bomb"],
            *bomb["release_point_m"], *bomb["explosion_point_m"], bomb["individual_duration_s"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
    workbook.save(target)


def write_q4() -> None:
    data = json.loads((OUTPUT / "q4.json").read_text(encoding="utf-8"))
    target, workbook, sheet = copy_template("result2.xlsx")
    for row, record in enumerate(data["records"], 2):
        values = [
            record["drone"], record["heading_deg"], record["speed_m_s"],
            *record["release_point_m"], *record["explosion_point_m"], record["individual_duration_s"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
    workbook.save(target)


def write_q5() -> None:
    data = json.loads((OUTPUT / "q5.json").read_text(encoding="utf-8"))
    target, workbook, sheet = copy_template("result3.xlsx")
    by_key = {(record["drone"], record["bomb"]): record for record in data["records"]}
    for row in range(2, 17):
        key = (sheet.cell(row, 1).value, sheet.cell(row, 4).value)
        record = by_key.get(key)
        if record is None:
            continue
        values = [
            record["drone"], record["heading_deg"], record["speed_m_s"], record["bomb"],
            *record["release_point_m"], *record["explosion_point_m"],
            duration(record["assigned_target_intervals_s"]), record["assigned_missile"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
    workbook.save(target)


def validate() -> None:
    q3 = json.loads((OUTPUT / "q3.json").read_text(encoding="utf-8"))
    q4 = json.loads((OUTPUT / "q4.json").read_text(encoding="utf-8"))
    q5 = json.loads((OUTPUT / "q5.json").read_text(encoding="utf-8"))
    validate_q5_constraints(q5["records"])
    expected = {
        "result1.xlsx": (3, 10),
        "result2.xlsx": (3, 10),
        "result3.xlsx": (len(q5["records"]), 12),
    }
    for name, (filled_rows, columns) in expected.items():
        sheet = openpyxl.load_workbook(OUTPUT / name, data_only=False).active
        rows = [row for row in range(2, sheet.max_row + 1) if sheet.cell(row, columns).value is not None]
        if len(rows) != filled_rows:
            raise RuntimeError(f"{name}: expected {filled_rows} filled result rows, found {len(rows)}")

    result1 = openpyxl.load_workbook(OUTPUT / "result1.xlsx", data_only=False).active
    for row, bomb in enumerate(q3["bombs"], 2):
        expected_values = [
            q3["heading_deg"], q3["speed_m_s"], bomb["bomb"], *bomb["release_point_m"],
            *bomb["explosion_point_m"], bomb["individual_duration_s"],
        ]
        assert_row([result1.cell(row, column).value for column in range(1, 11)], expected_values)

    result2 = openpyxl.load_workbook(OUTPUT / "result2.xlsx", data_only=False).active
    for row, record in enumerate(q4["records"], 2):
        expected_values = [
            record["drone"], record["heading_deg"], record["speed_m_s"],
            *record["release_point_m"], *record["explosion_point_m"], record["individual_duration_s"],
        ]
        assert_row([result2.cell(row, column).value for column in range(1, 11)], expected_values)

    result3 = openpyxl.load_workbook(OUTPUT / "result3.xlsx", data_only=False).active
    q5_by_key = {(record["drone"], record["bomb"]): record for record in q5["records"]}
    for row in range(2, 17):
        key = (result3.cell(row, 1).value, result3.cell(row, 4).value)
        record = q5_by_key.get(key)
        if record is None:
            assert all(result3.cell(row, column).value is None for column in range(2, 4))
            assert all(result3.cell(row, column).value is None for column in range(5, 13))
            continue
        expected_values = [
            record["drone"], record["heading_deg"], record["speed_m_s"], record["bomb"],
            *record["release_point_m"], *record["explosion_point_m"],
            duration(record["assigned_target_intervals_s"]), record["assigned_missile"],
        ]
        assert_row([result3.cell(row, column).value for column in range(1, 13)], expected_values)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--tier", choices=("working", "frozen"), default="working")
    args = parser.parse_args()
    OUTPUT = PROJECT_ROOT / "results" / args.tier
    write_q3()
    write_q4()
    write_q5()
    validate()
    print(f"{args.tier} result1.xlsx, result2.xlsx, and result3.xlsx written and field-validated")
